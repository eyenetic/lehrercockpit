from __future__ import annotations

from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
import json
import os
from pathlib import Path
import queue
import threading
import time
from urllib.parse import urlparse

try:
    from backend.dashboard import build_dashboard_payload
    from backend.local_settings import save_itslearning_settings
    from backend.classwork_scraper import scrape_classwork_plan
    from backend.classwork_cache import load_cache, save_cache, cache_age_minutes, get_previous_hash
    _DASHBOARD_IMPORT_ERROR: Exception | None = None
except Exception as _exc:
    build_dashboard_payload = None  # type: ignore[assignment]
    save_itslearning_settings = None  # type: ignore[assignment]
    scrape_classwork_plan = None  # type: ignore[assignment]
    load_cache = None  # type: ignore[assignment]
    save_cache = None  # type: ignore[assignment]
    cache_age_minutes = None  # type: ignore[assignment]
    get_previous_hash = None  # type: ignore[assignment]
    _DASHBOARD_IMPORT_ERROR = _exc
    import traceback
    traceback.print_exc()


PROJECT_ROOT = Path(__file__).resolve().parent
MOCK_DATA_PATH = PROJECT_ROOT / "data" / "mock-dashboard.json"
MONITOR_STATE_PATH = PROJECT_ROOT / "data" / "document-monitor-state.json"
CLASSWORK_CACHE_PATH = PROJECT_ROOT / "data" / "classwork-cache.json"
CLASSWORK_COOKIE_PATH = PROJECT_ROOT / "data" / "classwork-session.json"
ENV_FILE_PATH = PROJECT_ROOT / ".env.local"

CORS_ORIGIN = os.environ.get("CORS_ORIGIN", "*")


def _load_env_file() -> None:
    """Load .env.local into os.environ (only sets keys that aren't already set)."""
    env_path = PROJECT_ROOT / ".env.local"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'").strip('"')
        if key and key not in os.environ:
            os.environ[key] = value


_load_env_file()

CLASSWORK_PLAN_URL = os.environ.get("CLASSWORK_PLAN_URL", "")
MS_LOGIN_EMAIL = os.environ.get("MS_LOGIN_EMAIL", "")
MS_LOGIN_PASSWORD = os.environ.get("MS_LOGIN_PASSWORD", "")

# Cron: default every 6h, but also schedule at 05:00
SCRAPE_INTERVAL_SECONDS = int(os.environ.get("CLASSWORK_SCRAPE_INTERVAL_SECONDS", str(6 * 3600)))
SCRAPE_MIN_AGE_MINUTES = float(os.environ.get("CLASSWORK_SCRAPE_MIN_AGE_MINUTES", "30"))

# Global scrape state
_scrape_lock = threading.Lock()
_scrape_in_progress = False

# SSE progress queues: maps client_id -> Queue
_progress_queues: dict[str, queue.Queue] = {}
_progress_queues_lock = threading.Lock()


# ── Scrape runner ─────────────────────────────────────────────────────────────

def _broadcast_progress(percent: int, message: str) -> None:
    """Send progress to all connected SSE clients."""
    event = json.dumps({"percent": percent, "message": message})
    with _progress_queues_lock:
        dead = []
        for client_id, q in _progress_queues.items():
            try:
                q.put_nowait(event)
            except queue.Full:
                dead.append(client_id)
        for cid in dead:
            _progress_queues.pop(cid, None)


def _run_scrape(url: str, email: str = "", password: str = "") -> dict:
    """Run the scraper with progress broadcasting, save to cache, return result."""
    global _scrape_in_progress

    if not scrape_classwork_plan or not save_cache:
        return {"status": "error", "detail": "Scraper module not available."}

    with _scrape_lock:
        if _scrape_in_progress:
            cached = load_cache(CLASSWORK_CACHE_PATH) if load_cache else {}
            cached["scrapeInProgress"] = True
            cached["detail"] = "Scrape läuft bereits. Bitte warten."
            return cached
        _scrape_in_progress = True

    try:
        previous_hash = get_previous_hash(CLASSWORK_CACHE_PATH) if get_previous_hash else ""

        result = scrape_classwork_plan(
            url=url,
            email=email,
            password=password,
            cookie_path=CLASSWORK_COOKIE_PATH,
            previous_hash=previous_hash,
            on_progress=_broadcast_progress,
        )
        save_cache(CLASSWORK_CACHE_PATH, result)
        # Signal done to SSE clients
        _broadcast_progress(100, result.get("detail", "Fertig."))
        return result
    finally:
        _scrape_in_progress = False


def _seconds_until_five_am() -> float:
    """Return seconds until next 05:00 local time."""
    now = time.localtime()
    target_hour = 5
    seconds_today = now.tm_hour * 3600 + now.tm_min * 60 + now.tm_sec
    target_seconds = target_hour * 3600
    if seconds_today < target_seconds:
        return target_seconds - seconds_today
    else:
        return (24 * 3600) - seconds_today + target_seconds


def _cron_worker() -> None:
    """Background thread: runs scrape at 05:00 every day + interval-based fallback."""
    if not CLASSWORK_PLAN_URL:
        print("[classwork-cron] No CLASSWORK_PLAN_URL, cron disabled.", flush=True)
        return

    # Initial delay
    time.sleep(30)

    while True:
        try:
            # Schedule at 05:00
            wait = _seconds_until_five_am()
            if wait > SCRAPE_INTERVAL_SECONDS:
                # Use interval if 05:00 is too far away
                wait = SCRAPE_INTERVAL_SECONDS

            print(f"[classwork-cron] Next scrape in {wait/3600:.2f}h", flush=True)
            time.sleep(wait)

            age = cache_age_minutes(CLASSWORK_CACHE_PATH) if cache_age_minutes else None
            if age is None or age >= SCRAPE_MIN_AGE_MINUTES:
                print("[classwork-cron] Running scheduled scrape…", flush=True)
                _run_scrape(CLASSWORK_PLAN_URL, MS_LOGIN_EMAIL, MS_LOGIN_PASSWORD)
            else:
                print(f"[classwork-cron] Cache fresh ({age:.1f}min), skipping.", flush=True)

        except Exception as exc:
            print(f"[classwork-cron] Error: {exc}", flush=True)
            time.sleep(60)


# ── HTTP handler ──────────────────────────────────────────────────────────────

class LehrerCockpitHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(PROJECT_ROOT), **kwargs)

    def do_OPTIONS(self) -> None:
        self.send_response(HTTPStatus.NO_CONTENT)
        self._send_cors_headers()
        self.end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)

        if parsed.path == "/api/health":
            self._send_json({"status": "ok"})
            return

        if parsed.path == "/api/dashboard":
            if _DASHBOARD_IMPORT_ERROR is not None:
                self._send_json(
                    {"error": "dashboard module failed to import", "detail": str(_DASHBOARD_IMPORT_ERROR)},
                    status=HTTPStatus.INTERNAL_SERVER_ERROR,
                )
                return
            payload = build_dashboard_payload(MOCK_DATA_PATH, MONITOR_STATE_PATH, CLASSWORK_CACHE_PATH)
            self._send_json(payload)
            return

        if parsed.path == "/api/classwork":
            if load_cache is None:
                self._send_json({"status": "error", "detail": "Cache module not available."}, status=HTTPStatus.INTERNAL_SERVER_ERROR)
                return
            result = load_cache(CLASSWORK_CACHE_PATH)
            age = cache_age_minutes(CLASSWORK_CACHE_PATH) if cache_age_minutes else None
            result["cacheAgeMinutes"] = round(age, 1) if age is not None else None
            result["scrapeInProgress"] = _scrape_in_progress
            self._send_json(result)
            return

        if parsed.path == "/api/classwork/scrape/progress":
            # Server-Sent Events stream for live progress
            self._handle_sse_progress()
            return

        super().do_GET()

    def do_POST(self) -> None:
        parsed = urlparse(self.path)

        if parsed.path == "/api/classwork/upload":
            self._handle_classwork_upload()
            return

        if parsed.path == "/api/classwork/scrape":
            url = CLASSWORK_PLAN_URL
            if not url:
                self._send_json(
                    {"error": "no-url", "detail": "CLASSWORK_PLAN_URL ist nicht konfiguriert."},
                    status=HTTPStatus.BAD_REQUEST,
                )
                return

            if scrape_classwork_plan is None:
                self._send_json(
                    {"error": "no-scraper", "detail": "Scraper-Modul nicht verfuegbar."},
                    status=HTTPStatus.INTERNAL_SERVER_ERROR,
                )
                return

            if _scrape_in_progress:
                result = load_cache(CLASSWORK_CACHE_PATH) if load_cache else {}
                result["scrapeInProgress"] = True
                result["detail"] = "Scrape läuft bereits. Bitte warten."
                self._send_json(result)
                return

            thread = threading.Thread(
                target=_run_scrape,
                args=(url, MS_LOGIN_EMAIL, MS_LOGIN_PASSWORD),
                daemon=True,
            )
            thread.start()

            self._send_json(
                {"status": "started", "detail": "Scrape gestartet. Fortschritt via /api/classwork/scrape/progress abrufbar."},
                status=HTTPStatus.ACCEPTED,
            )
            return

        if parsed.path == "/api/local-settings/itslearning":
            if not self._is_local_request():
                self._send_json({"error": "local-only"}, status=HTTPStatus.FORBIDDEN)
                return

            if _DASHBOARD_IMPORT_ERROR is not None or save_itslearning_settings is None:
                self._send_json(
                    {"error": "settings module failed to import", "detail": str(_DASHBOARD_IMPORT_ERROR)},
                    status=HTTPStatus.INTERNAL_SERVER_ERROR,
                )
                return

            payload = self._read_json_body()
            username = str(payload.get("username", "")).strip()
            password = str(payload.get("password", "")).strip()
            base_url = str(payload.get("baseUrl", "https://berlin.itslearning.com")).strip() or "https://berlin.itslearning.com"
            max_updates = payload.get("maxUpdates", 6)

            if not username or not password:
                self._send_json(
                    {"error": "validation", "detail": "Benutzername und Passwort werden benoetigt."},
                    status=HTTPStatus.BAD_REQUEST,
                )
                return

            try:
                save_itslearning_settings(
                    ENV_FILE_PATH,
                    base_url=base_url,
                    username=username,
                    password=password,
                    max_updates=int(max_updates),
                )
            except Exception as exc:
                self._send_json(
                    {"error": "save-failed", "detail": f"{type(exc).__name__}: {exc}"},
                    status=HTTPStatus.INTERNAL_SERVER_ERROR,
                )
                return

            self._send_json({
                "status": "ok",
                "detail": "itslearning-Zugang lokal gespeichert.",
                "username": username,
                "baseUrl": base_url,
            })
            return

        self._send_json({"error": "not-found"}, status=HTTPStatus.NOT_FOUND)

    def _handle_classwork_upload(self) -> None:
        """Handle multipart file upload of XLS/XLSX, parse with openpyxl, save to cache."""
        content_type = self.headers.get("Content-Type", "")
        content_length = int(self.headers.get("Content-Length", "0") or 0)

        if not content_length:
            self._send_json({"error": "no-body", "detail": "Kein Dateiinhalt empfangen."}, status=HTTPStatus.BAD_REQUEST)
            return

        if content_length > 20 * 1024 * 1024:  # 20 MB max
            self._send_json({"error": "too-large", "detail": "Datei zu groß (max 20 MB)."}, status=HTTPStatus.REQUEST_ENTITY_TOO_LARGE)
            return

        raw_body = self.rfile.read(content_length)

        # Extract file bytes from multipart/form-data
        file_bytes = _extract_multipart_file(raw_body, content_type)
        if file_bytes is None:
            # Fallback: treat entire body as raw file bytes (e.g. direct PUT)
            file_bytes = raw_body

        try:
            result = _parse_classwork_xlsx(file_bytes)
        except Exception as exc:
            self._send_json(
                {"error": "parse-failed", "detail": f"Datei konnte nicht gelesen werden: {type(exc).__name__}: {exc}"},
                status=HTTPStatus.UNPROCESSABLE_ENTITY,
            )
            return

        if save_cache:
            save_cache(CLASSWORK_CACHE_PATH, result)

        _broadcast_progress(100, result.get("detail", "Hochgeladen."))
        self._send_json(result)

    def _handle_sse_progress(self) -> None:
        """Send Server-Sent Events for scrape progress."""
        import uuid
        client_id = str(uuid.uuid4())
        q: queue.Queue = queue.Queue(maxsize=50)

        with _progress_queues_lock:
            _progress_queues[client_id] = q

        try:
            self.send_response(200)
            self.send_header("Content-Type", "text/event-stream")
            self.send_header("Cache-Control", "no-cache")
            self.send_header("Connection", "keep-alive")
            self.send_header("Access-Control-Allow-Origin", CORS_ORIGIN)
            self.end_headers()

            # Send current scrape state immediately
            if _scrape_in_progress:
                self._write_sse({"percent": 0, "message": "Scrape läuft…"})
            else:
                cached = load_cache(CLASSWORK_CACHE_PATH) if load_cache else {}
                pct = 100 if cached.get("status") in ("ok", "warning") else 0
                msg = cached.get("detail", "Bereit.") or "Bereit."
                self._write_sse({"percent": pct, "message": msg[:100]})

            # Stream progress events
            while True:
                try:
                    event_data = q.get(timeout=25)
                    data = json.loads(event_data)
                    self._write_sse(data)
                    if data.get("percent", 0) >= 100:
                        break
                except queue.Empty:
                    # Send keepalive comment
                    self.wfile.write(b": keepalive\n\n")
                    self.wfile.flush()
                except (BrokenPipeError, ConnectionResetError):
                    break

        finally:
            with _progress_queues_lock:
                _progress_queues.pop(client_id, None)

    def _write_sse(self, data: dict) -> None:
        try:
            line = f"data: {json.dumps(data)}\n\n".encode("utf-8")
            self.wfile.write(line)
            self.wfile.flush()
        except Exception:
            pass

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store")
        self._send_cors_headers()
        super().end_headers()

    def _send_cors_headers(self) -> None:
        self.send_header("Access-Control-Allow-Origin", CORS_ORIGIN)
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def _send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=True).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_json_body(self) -> dict:
        content_length = int(self.headers.get("Content-Length", "0") or 0)
        raw_body = self.rfile.read(content_length) if content_length else b"{}"
        if not raw_body:
            return {}
        try:
            return json.loads(raw_body.decode("utf-8"))
        except json.JSONDecodeError:
            return {}

    def _is_local_request(self) -> bool:
        host, *_rest = self.client_address
        return host in {"127.0.0.1", "::1", "localhost"}

    def log_message(self, fmt: str, *args) -> None:
        # Suppress SSE keepalive log spam
        if "/api/classwork/scrape/progress" not in str(args):
            super().log_message(fmt, *args)


def _extract_multipart_file(body: bytes, content_type: str) -> bytes | None:
    """Extract raw file bytes from a multipart/form-data body."""
    import re as _re
    boundary_match = _re.search(r"boundary=([^\s;]+)", content_type)
    if not boundary_match:
        return None
    boundary = ("--" + boundary_match.group(1)).encode()
    parts = body.split(boundary)
    for part in parts:
        if b"filename=" not in part:
            continue
        # Skip headers, grab body after double CRLF
        sep = b"\r\n\r\n"
        idx = part.find(sep)
        if idx == -1:
            sep = b"\n\n"
            idx = part.find(sep)
        if idx == -1:
            continue
        file_data = part[idx + len(sep):]
        # Strip trailing boundary marker
        if file_data.endswith(b"\r\n"):
            file_data = file_data[:-2]
        return file_data
    return None


def _parse_classwork_xlsx(file_bytes: bytes) -> dict:
    """Parse XLS/XLSX bytes with openpyxl and return classwork cache dict."""
    from io import BytesIO as _BytesIO
    from datetime import datetime as _dt
    import hashlib as _hashlib

    now = _dt.now()

    try:
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(_BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        # Try as CSV fallback
        import csv as _csv
        text = file_bytes.decode("utf-8", errors="replace")
        reader = _csv.reader(text.splitlines())
        all_rows = [r for r in reader if any(c.strip() for c in r)]
        if not all_rows:
            raise ValueError("Datei ist leer oder kein lesbares Format.")
        header = [c.strip() for c in all_rows[0]]
        structured = [{header[i]: (row[i].strip() if i < len(row) else "") for i in range(len(header))} for row in all_rows[1:] if any(c.strip() for c in row)]
        preview = [" | ".join(c.strip() for c in row[:6] if c.strip()) for row in all_rows[:9]]
        return {
            "status": "ok", "title": "Klassenarbeitsplan",
            "detail": f"CSV hochgeladen. {len(structured)} Eintraege gelesen.",
            "updatedAt": now.strftime("%H:%M"), "scrapedAt": now.isoformat(),
            "previewRows": preview, "structuredRows": structured[:80],
            "sourceUrl": "", "scrapeMode": "upload",
            "dataHash": _hashlib.sha256(file_bytes).hexdigest()[:16],
            "hasChanges": False, "noChanges": False,
        }

    def _clean_header(value: str) -> str:
        """Normalize cell header: replace newlines/extra spaces with single space."""
        import re as _re
        return _re.sub(r"[\r\n]+", " ", str(value)).strip()

    all_structured: list[dict] = []
    preview_rows: list[str] = []
    total_sheets = len(wb.sheetnames)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        raw_rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(v).strip() if v is not None else "" for v in row]
            if any(v for v in values):
                raw_rows.append(values)
            if len(raw_rows) >= 60:
                break

        if not raw_rows:
            continue

        header = [_clean_header(col) if col else f"Spalte{i+1}" for i, col in enumerate(raw_rows[0])]

        for row in raw_rows[1:]:
            if not any(v for v in row):
                continue
            entry: dict = {"_sheet": sheet_name}
            for i, col in enumerate(header):
                entry[col] = row[i] if i < len(row) else ""
            all_structured.append(entry)

        # Build preview from first sheet
        if not preview_rows:
            preview_rows = [" | ".join(v for v in row[:6] if v) for row in raw_rows[:9] if any(row)]

    wb.close()

    if not all_structured:
        raise ValueError("Tabelle ist leer oder kein lesbares Format.")

    total_entries = len(all_structured)
    return {
        "status": "ok", "title": "Klassenarbeitsplan",
        "detail": f"Excel-Datei hochgeladen. {total_entries} Eintraege aus {total_sheets} Tabellenblättern gelesen.",
        "updatedAt": now.strftime("%H:%M"), "scrapedAt": now.isoformat(),
        "previewRows": preview_rows, "structuredRows": all_structured[:200],
        "sourceUrl": "", "scrapeMode": "upload",
        "dataHash": _hashlib.sha256(file_bytes).hexdigest()[:16],
        "hasChanges": False, "noChanges": False,
    }


def run() -> None:
    port = int(os.environ.get("PORT", 8080))
    host = "0.0.0.0"

    if CLASSWORK_PLAN_URL:
        cron_thread = threading.Thread(target=_cron_worker, daemon=True)
        cron_thread.start()
        print(f"[classwork-cron] Cron started. Daily at 05:00 + every {SCRAPE_INTERVAL_SECONDS/3600:.1f}h", flush=True)
    else:
        print("[classwork-cron] CLASSWORK_PLAN_URL not set, cron disabled.", flush=True)

    server = ThreadingHTTPServer((host, port), LehrerCockpitHandler)
    print(f"Lehrer-Cockpit läuft auf http://{host}:{port}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    run()
