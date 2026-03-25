"""Shared file-parsing utilities used by both app.py (Flask/Render) and server.py (local dev).

Moving these here eliminates the verbatim duplication that existed in both entry-point files.
"""
from __future__ import annotations

import hashlib
import re as _re
from typing import Optional


def extract_multipart_file(body: bytes, content_type: str) -> Optional[bytes]:
    """Extract raw file bytes from a multipart/form-data body.

    Returns the raw bytes of the first part that has a ``filename=`` header,
    or ``None`` if no such part is found or the body is not multipart.
    """
    boundary_match = _re.search(r"boundary=([^\s;]+)", content_type)
    if not boundary_match:
        return None
    boundary = ("--" + boundary_match.group(1)).encode()
    parts = body.split(boundary)
    for part in parts:
        if b"filename=" not in part:
            continue
        sep = b"\r\n\r\n"
        idx = part.find(sep)
        if idx == -1:
            sep = b"\n\n"
            idx = part.find(sep)
        if idx == -1:
            continue
        file_data = part[idx + len(sep):]
        if file_data.endswith(b"\r\n"):
            file_data = file_data[:-2]
        return file_data
    return None


def parse_classwork_xlsx(file_bytes: bytes) -> dict:
    """Parse XLS/XLSX bytes (or fall back to CSV) and return a classwork cache dict.

    Raises ``ValueError`` if the file is empty or unreadable after all fallbacks.
    """
    from io import BytesIO as _BytesIO
    from datetime import datetime as _dt

    now = _dt.now()

    def _clean_header(value: str) -> str:
        return _re.sub(r"[\r\n]+", " ", str(value)).strip()

    try:
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(_BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        # CSV fallback
        import csv as _csv
        text = file_bytes.decode("utf-8", errors="replace")
        reader = _csv.reader(text.splitlines())
        all_rows = [r for r in reader if any(c.strip() for c in r)]
        if not all_rows:
            raise ValueError("Datei ist leer oder kein lesbares Format.")
        header = [c.strip() for c in all_rows[0]]
        structured = [
            {header[i]: (row[i].strip() if i < len(row) else "") for i in range(len(header))}
            for row in all_rows[1:] if any(c.strip() for c in row)
        ]
        preview = [" | ".join(c.strip() for c in row[:6] if c.strip()) for row in all_rows[:9]]
        return {
            "status": "ok",
            "title": "Klassenarbeitsplan",
            "detail": f"CSV hochgeladen. {len(structured)} Eintraege gelesen.",
            "updatedAt": now.strftime("%H:%M"),
            "scrapedAt": now.isoformat(),
            "previewRows": preview,
            "structuredRows": structured[:200],
            "sourceUrl": "",
            "scrapeMode": "upload",
            "dataHash": hashlib.sha256(file_bytes).hexdigest()[:16],
            "hasChanges": False,
            "noChanges": False,
        }

    all_structured: list[dict] = []
    preview_rows: list[str] = []
    total_sheets = len(wb.sheetnames)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        raw_rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(v).strip() if v is not None else "" for v in row]
            if any(v for v in values):
                raw_rows.append(values)
            if len(raw_rows) >= 60:
                break

        if not raw_rows:
            continue

        header = [_clean_header(col) if col else f"Spalte{i+1}" for i, col in enumerate(raw_rows[0])]

        for row in raw_rows[1:]:
            if not any(v for v in row):
                continue
            entry: dict = {"_sheet": sheet_name}
            for i, col in enumerate(header):
                entry[col] = row[i] if i < len(row) else ""
            all_structured.append(entry)

        if not preview_rows:
            preview_rows = [" | ".join(v for v in row[:6] if v) for row in raw_rows[:9] if any(row)]

    wb.close()

    if not all_structured:
        raise ValueError("Tabelle ist leer oder kein lesbares Format.")

    return {
        "status": "ok",
        "title": "Klassenarbeitsplan",
        "detail": f"Excel-Datei hochgeladen. {len(all_structured)} Eintraege aus {total_sheets} Tabellenblättern gelesen.",
        "updatedAt": now.strftime("%H:%M"),
        "scrapedAt": now.isoformat(),
        "previewRows": preview_rows,
        "structuredRows": all_structured[:200],
        "sourceUrl": "",
        "scrapeMode": "upload",
        "dataHash": hashlib.sha256(file_bytes).hexdigest()[:16],
        "hasChanges": False,
        "noChanges": False,
    }
