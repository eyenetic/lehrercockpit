"""Flask WSGI entry point for Render / gunicorn.

All API logic is ported from server.py.  Static files (index.html, styles.css,
src/, icons/, data/, manifest.json) are served by Flask's send_from_directory
so the app works as a single-process web service without a separate static host.
"""
from __future__ import annotations

import hashlib
import json
import os
from pathlib import Path

from flask import Flask, Response, jsonify, request, send_from_directory

# ── Load .env.local early ─────────────────────────────────────────────────────

PROJECT_ROOT = Path(__file__).resolve().parent


def _load_env_file() -> None:
    env_path = PROJECT_ROOT / ".env.local"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'").strip('"')
        if key and key not in os.environ:
            os.environ[key] = value


_load_env_file()

# ── Backend imports ───────────────────────────────────────────────────────────

try:
    from backend.dashboard import build_dashboard_payload
    from backend.classwork_cache import load_cache, save_cache
    from backend.grades_store import create_grade_entry, load_gradebook, save_gradebook
    from backend.notes_store import create_note, load_notes, save_notes
    from backend.local_settings import save_classwork_file, save_itslearning_settings
    _IMPORT_ERROR: Exception | None = None
except Exception as _exc:
    build_dashboard_payload = None  # type: ignore[assignment]
    save_classwork_file = None  # type: ignore[assignment]
    save_itslearning_settings = None  # type: ignore[assignment]
    load_cache = None  # type: ignore[assignment]
    save_cache = None  # type: ignore[assignment]
    create_grade_entry = None  # type: ignore[assignment]
    load_gradebook = None  # type: ignore[assignment]
    save_gradebook = None  # type: ignore[assignment]
    create_note = None  # type: ignore[assignment]
    load_notes = None  # type: ignore[assignment]
    save_notes = None  # type: ignore[assignment]
    _IMPORT_ERROR = _exc
    import traceback
    traceback.print_exc()

# ── Path constants ────────────────────────────────────────────────────────────

MOCK_DATA_PATH = PROJECT_ROOT / "data" / "mock-dashboard.json"
MONITOR_STATE_PATH = PROJECT_ROOT / "data" / "document-monitor-state.json"
CLASSWORK_CACHE_PATH = PROJECT_ROOT / "data" / "classwork-cache.json"
WEBUNTIS_CACHE_PATH = PROJECT_ROOT / "data" / "webuntis-cache.json"
GRADES_LOCAL_PATH = PROJECT_ROOT / "data" / "grades-local.json"
NOTES_LOCAL_PATH = PROJECT_ROOT / "data" / "class-notes-local.json"
ENV_FILE_PATH = PROJECT_ROOT / ".env.local"
CLASSWORK_LOCAL_PATH = PROJECT_ROOT / "data" / "classwork-plan-local.xlsx"

# ── Flask app ─────────────────────────────────────────────────────────────────

app = Flask(__name__, static_folder=None)
app.config["PROPAGATE_EXCEPTIONS"] = False

CORS_ORIGIN = os.environ.get("CORS_ORIGIN", "*")


def _cors(response: Response) -> Response:
    response.headers["Access-Control-Allow-Origin"] = CORS_ORIGIN
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type"
    response.headers["Cache-Control"] = "no-store"
    return response


@app.after_request
def after_request(response: Response) -> Response:
    return _cors(response)


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def static_files(path: str) -> Response:
    """Serve all static files from PROJECT_ROOT."""
    if not path:
        path = "index.html"
    # Security: never serve Python source or hidden files via this route
    blocked_exts = {".py", ".pyc", ".env"}
    blocked_names = {".env.local", ".gitignore"}
    if Path(path).suffix in blocked_exts or Path(path).name in blocked_names:
        return jsonify({"error": "forbidden"}), 403
    file_path = PROJECT_ROOT / path
    if file_path.is_dir():
        file_path = file_path / "index.html"
        path = str(file_path.relative_to(PROJECT_ROOT))
    return send_from_directory(str(PROJECT_ROOT), path)


# ── GET endpoints ─────────────────────────────────────────────────────────────

@app.route("/api/health")
def api_health() -> Response:
    return jsonify({"status": "ok"})


@app.route("/api/dashboard")
def api_dashboard() -> Response:
    if _IMPORT_ERROR is not None:
        return jsonify({"error": "dashboard module failed to import", "detail": str(_IMPORT_ERROR)}), 500
    payload = build_dashboard_payload(
        MOCK_DATA_PATH,
        MONITOR_STATE_PATH,
        CLASSWORK_CACHE_PATH,
        WEBUNTIS_CACHE_PATH,
    )
    return jsonify(payload)


@app.route("/api/classwork")
def api_classwork() -> Response:
    if load_cache is None:
        return jsonify({"status": "error", "detail": "Cache module not available."}), 500
    return jsonify(load_cache(CLASSWORK_CACHE_PATH))


@app.route("/api/grades")
def api_grades() -> Response:
    if load_gradebook is None:
        return jsonify({"status": "error", "detail": "Grades store not available."}), 500
    return jsonify(load_gradebook(GRADES_LOCAL_PATH))


@app.route("/api/notes")
def api_notes() -> Response:
    if load_notes is None:
        return jsonify({"status": "error", "detail": "Notes store not available."}), 500
    return jsonify(load_notes(NOTES_LOCAL_PATH))


# ── POST endpoints ────────────────────────────────────────────────────────────

@app.route("/api/classwork/upload", methods=["POST", "OPTIONS"])
def api_classwork_upload() -> Response:
    if request.method == "OPTIONS":
        return Response(status=204)

    content_length = request.content_length or 0
    if not content_length:
        return jsonify({"error": "no-body", "detail": "Kein Dateiinhalt empfangen."}), 400
    if content_length > 20 * 1024 * 1024:
        return jsonify({"error": "too-large", "detail": "Datei zu groß (max 20 MB)."}), 413

    raw_body = request.get_data()
    content_type = request.content_type or ""
    file_bytes = _extract_multipart_file(raw_body, content_type)
    if file_bytes is None:
        file_bytes = raw_body

    try:
        result = _parse_classwork_xlsx(file_bytes)
    except Exception as exc:
        return jsonify({"error": "parse-failed", "detail": f"Datei konnte nicht gelesen werden: {type(exc).__name__}: {exc}"}), 422

    if save_cache:
        save_cache(CLASSWORK_CACHE_PATH, result)

    return jsonify(result)


@app.route("/api/classwork/browser-fetch", methods=["POST", "OPTIONS"])
def api_classwork_browser_fetch() -> Response:
    if request.method == "OPTIONS":
        return Response(status=204)

    import socket
    hostname = socket.gethostname()
    is_local = (
        os.environ.get("IS_LOCAL_RUNTIME", "").lower() in ("1", "true", "yes")
        or "render" not in hostname.lower()
    )
    if not is_local:
        return jsonify({"status": "error", "detail": "Browser-Abruf ist nur lokal verfuegbar, nicht auf dem Server."}), 400

    body = request.get_json(silent=True) or {}
    onedrive_url = body.get("url", "") or os.environ.get("CLASSWORK_ONEDRIVE_URL", "")
    if not onedrive_url:
        return jsonify({"status": "error", "detail": "Kein OneDrive-Link konfiguriert. Bitte CLASSWORK_ONEDRIVE_URL in .env.local setzen."}), 400

    try:
        from backend.classwork_browser import fetch_classwork_from_browser, write_to_cache
    except ImportError as exc:
        return jsonify({"status": "error", "detail": f"classwork_browser.py fehlt: {exc}"}), 500

    result = fetch_classwork_from_browser(onedrive_url)
    if result.get("status") == "ok":
        result["cacheWritten"] = write_to_cache(result)
    return jsonify(result)


@app.route("/api/local-settings/itslearning", methods=["POST", "OPTIONS"])
def api_local_settings_itslearning() -> Response:
    if request.method == "OPTIONS":
        return Response(status=204)
    if not _is_local_request():
        return jsonify({"error": "local-only"}), 403
    if _IMPORT_ERROR is not None or save_itslearning_settings is None:
        return jsonify({"error": "settings module failed to import", "detail": str(_IMPORT_ERROR)}), 500

    payload = request.get_json(silent=True) or {}
    username = str(payload.get("username", "")).strip()
    password = str(payload.get("password", "")).strip()
    base_url = str(payload.get("baseUrl", "https://berlin.itslearning.com")).strip() or "https://berlin.itslearning.com"
    max_updates = payload.get("maxUpdates", 6)

    if not username or not password:
        return jsonify({"error": "validation", "detail": "Benutzername und Passwort werden benoetigt."}), 400

    try:
        save_itslearning_settings(
            ENV_FILE_PATH,
            base_url=base_url,
            username=username,
            password=password,
            max_updates=int(max_updates),
        )
    except Exception as exc:
        return jsonify({"error": "save-failed", "detail": f"{type(exc).__name__}: {exc}"}), 500

    return jsonify({"status": "ok", "detail": "itslearning-Zugang lokal gespeichert.", "username": username, "baseUrl": base_url})


@app.route("/api/local-settings/classwork-upload", methods=["POST", "OPTIONS"])
def api_local_settings_classwork_upload() -> Response:
    if request.method == "OPTIONS":
        return Response(status=204)
    if not _is_local_request():
        return jsonify({"error": "local-only"}), 403
    if _IMPORT_ERROR is not None or save_classwork_file is None:
        return jsonify({"error": "settings module failed to import", "detail": str(_IMPORT_ERROR)}), 500

    payload = request.get_json(silent=True) or {}
    filename = str(payload.get("filename", "")).strip()
    content_base64 = str(payload.get("contentBase64", "")).strip()

    if not filename or not content_base64:
        return jsonify({"error": "validation", "detail": "Bitte eine XLSX-Datei auswaehlen."}), 400

    try:
        save_classwork_file(CLASSWORK_LOCAL_PATH, filename=filename, content_base64=content_base64)
    except Exception as exc:
        return jsonify({"error": "save-failed", "detail": f"{type(exc).__name__}: {exc}"}), 400

    return jsonify({"status": "ok", "detail": "Klassenarbeitsplan lokal importiert. Das Cockpit liest die Datei jetzt neu ein.", "path": str(CLASSWORK_LOCAL_PATH)})


@app.route("/api/local-settings/grades", methods=["POST", "OPTIONS"])
def api_local_settings_grades() -> Response:
    if request.method == "OPTIONS":
        return Response(status=204)
    if not _is_local_request():
        return jsonify({"error": "local-only"}), 403
    if load_gradebook is None or save_gradebook is None or create_grade_entry is None:
        return jsonify({"error": "grades module failed to import", "detail": str(_IMPORT_ERROR)}), 500

    payload = request.get_json(silent=True) or {}
    mode = str(payload.get("mode", "append")).strip() or "append"
    current = load_gradebook(GRADES_LOCAL_PATH)
    current_entries = current.get("entries", [])

    if mode == "replace":
        entries = payload.get("entries", [])
        result = save_gradebook(GRADES_LOCAL_PATH, entries if isinstance(entries, list) else [])
        return jsonify({"status": "ok", "detail": "Noten lokal gespeichert.", **result})

    if mode == "delete":
        entry_id = str(payload.get("id", "")).strip()
        if not entry_id:
            return jsonify({"error": "validation", "detail": "Eintrags-ID fehlt."}), 400
        remaining = [e for e in current_entries if e.get("id") != entry_id]
        result = save_gradebook(GRADES_LOCAL_PATH, remaining)
        return jsonify({"status": "ok", "detail": "Eintrag entfernt.", **result})

    entry = create_grade_entry(payload)
    if not entry["classLabel"] or not entry["studentName"] or not entry["title"]:
        return jsonify({"error": "validation", "detail": "Klasse, Schueler:in und Titel werden benoetigt."}), 400
    result = save_gradebook(GRADES_LOCAL_PATH, [entry] + current_entries)
    return jsonify({"status": "ok", "detail": "Note lokal gespeichert.", **result})


@app.route("/api/local-settings/notes", methods=["POST", "OPTIONS"])
def api_local_settings_notes() -> Response:
    if request.method == "OPTIONS":
        return Response(status=204)
    if not _is_local_request():
        return jsonify({"error": "local-only"}), 403
    if load_notes is None or save_notes is None or create_note is None:
        return jsonify({"error": "notes module failed to import", "detail": str(_IMPORT_ERROR)}), 500

    payload = request.get_json(silent=True) or {}
    mode = str(payload.get("mode", "upsert")).strip() or "upsert"
    current = load_notes(NOTES_LOCAL_PATH)
    current_notes = current.get("notes", [])

    if mode == "delete":
        class_label = str(payload.get("classLabel", "")).strip().upper()
        if not class_label:
            return jsonify({"error": "validation", "detail": "Klasse fehlt."}), 400
        remaining = [item for item in current_notes if item.get("classLabel") != class_label]
        result = save_notes(NOTES_LOCAL_PATH, remaining)
        return jsonify({"status": "ok", "detail": "Notiz entfernt.", **result})

    note = create_note(payload)
    if not note["classLabel"]:
        return jsonify({"error": "validation", "detail": "Klasse wird benoetigt."}), 400
    remaining = [item for item in current_notes if item.get("classLabel") != note["classLabel"]]
    if note["text"]:
        remaining = [note] + remaining
    result = save_notes(NOTES_LOCAL_PATH, remaining)
    return jsonify({"status": "ok", "detail": "Notiz lokal gespeichert.", **result})


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_local_request() -> bool:
    remote = request.remote_addr or ""
    return remote in {"127.0.0.1", "::1", "localhost"}


def _extract_multipart_file(body: bytes, content_type: str) -> bytes | None:
    import re as _re
    boundary_match = _re.search(r"boundary=([^\s;]+)", content_type)
    if not boundary_match:
        return None
    boundary = ("--" + boundary_match.group(1)).encode()
    parts = body.split(boundary)
    for part in parts:
        if b"filename=" not in part:
            continue
        sep = b"\r\n\r\n"
        idx = part.find(sep)
        if idx == -1:
            sep = b"\n\n"
            idx = part.find(sep)
        if idx == -1:
            continue
        file_data = part[idx + len(sep):]
        if file_data.endswith(b"\r\n"):
            file_data = file_data[:-2]
        return file_data
    return None


def _parse_classwork_xlsx(file_bytes: bytes) -> dict:
    from io import BytesIO as _BytesIO
    from datetime import datetime as _dt
    import re as _re

    now = _dt.now()

    def _clean_header(value: str) -> str:
        return _re.sub(r"[\r\n]+", " ", str(value)).strip()

    try:
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(_BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        import csv as _csv
        text = file_bytes.decode("utf-8", errors="replace")
        reader = _csv.reader(text.splitlines())
        all_rows = [r for r in reader if any(c.strip() for c in r)]
        if not all_rows:
            raise ValueError("Datei ist leer oder kein lesbares Format.")
        header = [c.strip() for c in all_rows[0]]
        structured = [
            {header[i]: (row[i].strip() if i < len(row) else "") for i in range(len(header))}
            for row in all_rows[1:] if any(c.strip() for c in row)
        ]
        preview = [" | ".join(c.strip() for c in row[:6] if c.strip()) for row in all_rows[:9]]
        return {
            "status": "ok", "title": "Klassenarbeitsplan",
            "detail": f"CSV hochgeladen. {len(structured)} Eintraege gelesen.",
            "updatedAt": now.strftime("%H:%M"), "scrapedAt": now.isoformat(),
            "previewRows": preview, "structuredRows": structured[:200],
            "sourceUrl": "", "scrapeMode": "upload",
            "dataHash": hashlib.sha256(file_bytes).hexdigest()[:16],
            "hasChanges": False, "noChanges": False,
        }

    all_structured: list[dict] = []
    preview_rows: list[str] = []
    total_sheets = len(wb.sheetnames)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        raw_rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = [str(v).strip() if v is not None else "" for v in row]
            if any(v for v in values):
                raw_rows.append(values)
            if len(raw_rows) >= 60:
                break

        if not raw_rows:
            continue

        header = [_clean_header(col) if col else f"Spalte{i+1}" for i, col in enumerate(raw_rows[0])]

        for row in raw_rows[1:]:
            if not any(v for v in row):
                continue
            entry: dict = {"_sheet": sheet_name}
            for i, col in enumerate(header):
                entry[col] = row[i] if i < len(row) else ""
            all_structured.append(entry)

        if not preview_rows:
            preview_rows = [" | ".join(v for v in row[:6] if v) for row in raw_rows[:9] if any(row)]

    wb.close()

    if not all_structured:
        raise ValueError("Tabelle ist leer oder kein lesbares Format.")

    return {
        "status": "ok", "title": "Klassenarbeitsplan",
        "detail": f"Excel-Datei hochgeladen. {len(all_structured)} Eintraege aus {total_sheets} Tabellenblättern gelesen.",
        "updatedAt": now.strftime("%H:%M"), "scrapedAt": now.isoformat(),
        "previewRows": preview_rows, "structuredRows": all_structured[:200],
        "sourceUrl": "", "scrapeMode": "upload",
        "dataHash": hashlib.sha256(file_bytes).hexdigest()[:16],
        "hasChanges": False, "noChanges": False,
    }


# ── Dev entrypoint ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8080))
    print(f"Lehrer-Cockpit (Flask dev) läuft auf http://0.0.0.0:{port}", flush=True)
    app.run(host="0.0.0.0", port=port, debug=False)
